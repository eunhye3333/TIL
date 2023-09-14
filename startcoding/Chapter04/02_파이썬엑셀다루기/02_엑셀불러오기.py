import openpyxl

fpath = r'E:\develop\startcoding\Chapter04\02_파이썬엑셀다루기\출연자_data.xlsx'

# 1) 엑셀 불러오기
wb = openpyxl.load_workbook(fpath)

# 2) 엑셀 시트 선택
ws = wb['놀라운토요일']

# 3) 데이터 수정하기
ws['A3'] = 456
ws['B3'] = '키'

# 4) 엑셀 저장하기
wb.save(fpath)