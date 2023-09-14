# -*- coding: utf-8 -*-
import random

import openpyxl as xl
import cx_Oracle
import os
from random import *
import random
import string
import datetime
from datetime import datetime
from time import sleep

os.putenv('NLS_LANG', '.UTF8')

conn = cx_Oracle.connect('SHINWOO/SHINWOO@localhost:1521/xe')
# conn2 = cx_Oracle.connect('SHINWOO/SHINWOO@localhost:1521/xe')
cs = conn.cursor()
cs2 = conn.cursor()

wb = xl.load_workbook('C:/dev/book.xlsx')
string_pool = 'a', 'b'
for sheet_nm in wb.sheetnames:
    # print('*' * 100)
    # print('시트명:', sheet_nm)
    sheet = wb[sheet_nm]
    for row_data in sheet.iter_rows(min_row=2):
        # for cell in row_data:
        # for i in range(len(row_data)):
        #     print('[', row_data[i].value, ']')
        # print('=' * 100)
        i1 = randint(100, 999)
        i2 = randint(1000, 9999)
        s1 = ''.join(random.choice(string.ascii_letters))
        # sql = f"insert into BOOK values ('{str(row_data[1].value) + str(i)}', '{row_data[0].value}', {row_data[1].value}, '{row_data[2].value}', '{row_data[3].value}', SYSDATE, DEFAULT, DEFAULT, '{row_data[5].value}', NULL, '{row_data[4].value}')"
        # sql = "INSERT INTO BOOK VALUES(:1, :2, :3, :4, :5, SYSDATE, DEFAULT, DEFAULT, :6, NULL, :7)"
        sql = "INSERT INTO BOOK VALUES(:1, :2, :3, :4, :5, SYSDATE, DEFAULT, DEFAULT, :6, NULL, :7)"
        bookNo = str(row_data[1].value) + "." + str(i1) + s1 + str(i2)
        releDate = datetime.strptime(row_data[5].value, "%Y.%m.%d")
        val = (bookNo, row_data[0].value, int(row_data[1].value), row_data[2].value, row_data[3].value, row_data[4].value, releDate)
        cs.execute(sql, val)

        old_name = row_data[6].value
        new_name = str(round(datetime.utcnow().timestamp() * 1000)) + ".jpg"

        file_oldname = os.path.join("image", old_name)
        file_newname_newfile = os.path.join("image", new_name)
        os.rename(file_oldname, file_newname_newfile)

        sql2 = "INSERT INTO PHOTO VALUES(SEQ_PHOTO.NEXTVAL, :1, :2, :3, :4, DEFAULT, DEFAULT)"
        val2 = (bookNo, old_name, new_name, "\ShinwooUniversity\SemiProject\WebContent\image/")
        cs.execute(sql2, val2)

        # conn.commit()
        # print(sql)
print("완료")
# os.system('pause')
conn.commit()
# conn2.commit()
sleep(2)
wb.close()